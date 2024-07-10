import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import math
import openpyxl
from calc_functions import calculations_CSTR

class ExperimentApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Experiment data")
        self.tabControl = ttk.Notebook(self)
        style = ttk.Style()
        style.configure('TNotebook.Tab', font=("Helvetica", 12))
        style.configure('TFrame', background='#e0dcfd')
        
        self.setup_tab = ttk.Frame(self.tabControl)
        self.reagents_tab = ttk.Frame(self.tabControl)
        self.experiment_tab = ttk.Frame(self.tabControl)
        self.run_tab = ttk.Frame(self.tabControl)
        self.help_tab = ttk.Frame(self.tabControl)

        label_font = ("Helvetica", 12)
        entry_font = ("Helvetica", 12)
        button_style = {'bg': '#c7bffb', 'fg': 'black', 'font': ("Helvetica", 12)}

        self.tabControl.add(self.setup_tab, text='Setup')
        self.tabControl.add(self.reagents_tab, text='Reagents')
        self.tabControl.add(self.experiment_tab, text='Experiment')
        self.tabControl.add(self.run_tab, text='Run')
        self.tabControl.add(self.help_tab, text='Help')

        self.tabControl.pack(expand=1, fill="both")

        # Setup Configuration Page
        self.exp_name = tk.Label(self.setup_tab, text="Experiment name:", font=label_font, bg='#e0dcfd')
        self.exp_name_entry = tk.Entry(self.setup_tab, font=entry_font)
        self.exp_name.grid(row=0, column=0)
        self.exp_name_entry.grid(row=0, column=1)

        self.cstr_volume_label = tk.Label(self.setup_tab, text="CSTR volume (ml):", font=label_font, bg='#e0dcfd')
        self.cstr_volume_entry = tk.Entry(self.setup_tab, font=entry_font)
        self.cstr_volume_label.grid(row=1, column=0)
        self.cstr_volume_entry.grid(row=1, column=1)

        self.num_syringes_label = tk.Label(self.setup_tab, text="Number of syringes:", font=label_font, bg='#e0dcfd')
        self.num_syringes_combobox = ttk.Combobox(self.setup_tab, values=list(range(1, 11)), font=entry_font)
        self.num_syringes_label.grid(row=2, column=0)
        self.num_syringes_combobox.grid(row=2, column=1)

        save_button = tk.Button(self.setup_tab, text="Save", command=self.save_CSTR, bg = '#c7bffb', fg = 'black', font = ("Helvetica", 12))
        save_button.grid(row=3, columnspan=2)
        

        # Reagents Page
        self.num_reagents_label = tk.Label(self.reagents_tab, text="Number of reagents:", font=label_font, bg='#e0dcfd')
        self.num_reagents_combobox = ttk.Combobox(self.reagents_tab, values=list(range(1, 11)), font=entry_font)
        self.num_reagents_label.grid(row=0, column=0)
        self.num_reagents_combobox.grid(row=0, column=1)

        self.num_add_reagents_label = tk.Label(self.reagents_tab, text="Number of reagents with two different concentrations:", font=label_font, bg='#e0dcfd')
        self.num_add_reagents_combobox = ttk.Combobox(self.reagents_tab, values=list(range(0, 10)), font=entry_font)
        self.num_add_reagents_label.grid(row=1, column=0)
        self.num_add_reagents_combobox.grid(row=1, column=1)

        self.add_boxes_button = tk.Button(self.reagents_tab, text="Add reagents list", command=self.add_reagent_names, **button_style)
        self.add_boxes_button.grid(row=2, columnspan=2)

        # Experiment Page
        self.flows = tk.Label(self.experiment_tab, text="Enter flow rates (microliters in hour) and time (min) when they were changed", font=label_font, bg='#e0dcfd')
        self.flows.grid(row=0)

        self.total_time = tk.Label(self.experiment_tab, text="Total time of experiment (min):", font=label_font, bg='#e0dcfd')
        self.total_time_entry = tk.Entry(self.experiment_tab, font=entry_font)
        self.total_time.grid(row=1, column=0)
        self.total_time_entry.grid(row=1, column=1)

        add_rows_button = tk.Button(self.experiment_tab, text="Add change", command=self.add_rows, **button_style)
        add_rows_button.grid(row=2, column=0)
        save_changes_button = tk.Button(self.experiment_tab, text="Save changes", command=self.save_changes, **button_style)
        save_changes_button.grid(row=2, column=1)
        self.num_rows = 4

        # Run Page
        self.message = tk.Label(self.run_tab, text="Make sure that you filled all the gaps with the correct data.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=1)

        run_button = tk.Button(self.run_tab, text="Start calculations", command=self.calc, **button_style)
        run_button.grid(row=3)
        
        # Help Page
        self.message = tk.Label(self.help_tab, text="1. Fill out the tabs sequentially, then on the run tab click the button to start calculations.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=0)
        self.message = tk.Label(self.help_tab, text="2. You must click the save button after filling out each tab.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=1)
        self.message = tk.Label(self.help_tab, text="3. Click the buttons to add reagents only after you have filled out the rest of the information on the tab.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=2)
        self.message = tk.Label(self.help_tab, text="4. If you are dispensing one reagent from two syringes, enter it into the reagent table twice.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=3)
        self.message = tk.Label(self.help_tab, text="5. Click the add change button to add one new row to the table on the third tab. The number of rows is not limited.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=4)
        self.message = tk.Label(self.help_tab, text="6. After adding all changes in the experiment tab, click the save button.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=5)
        self.message = tk.Label(self.help_tab, text= "7. Use a period as a fraction separator.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=6)
        self.message = tk.Label(self.help_tab, text="8. After saving, you can view the entered data in an Excel file.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=7)
        self.message = tk.Label(self.help_tab, text="9. After performing the calculations, new tabs with results for each substance will appear in this file.", font=label_font, bg='#e0dcfd')
        self.message.grid(row=8)


    

    # GUI functions

    def save_CSTR(self):
        self.create_experiment()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "CSTR"
        data = {'CSTR volume': [float(self.cstr_volume_entry.get())*1000]}
        for key, value in data.items():
            worksheet.append([key, *value])
        filename = self.exp_name_entry.get() + '.xlsx'
        workbook.save(filename)

    def add_reagent_names(self):
        num_rows = int(self.num_reagents_combobox.get()) + int(self.num_add_reagents_combobox.get())
        num_syr = int(self.num_syringes_combobox.get())
        label_reagent_name = tk.Label(self.reagents_tab, text=f"Reagent name", font=("Helvetica", 12), bg='#e0dcfd')
        label_syringe_in = tk.Label(self.reagents_tab, text=f"Syringe", font=("Helvetica", 12), bg='#e0dcfd')
        label_concentration = tk.Label(self.reagents_tab, text=f"Concentration (mM)", font=("Helvetica", 12), bg='#e0dcfd')
        label_reagent_name.grid(row=2, column=1)
        label_syringe_in.grid(row=2, column=2)
        label_concentration.grid(row=2, column=3)

        for i in range(num_rows):
            label_reagent = tk.Label(self.reagents_tab, text=f"{i+1}", font=("Helvetica", 12), bg='#e0dcfd')
            entry_reagent_name = tk.Entry(self.reagents_tab, font=("Helvetica", 12))
            entry_syringe_in = ttk.Combobox(self.reagents_tab, values=list(range(1, int(num_syr)+1)), font=("Helvetica", 12))
            entry_conc = tk.Entry(self.reagents_tab, font=("Helvetica", 12))
            label_reagent.grid(row=i+3, column=0)
            entry_reagent_name.grid(row=i+3, column=1)
            entry_syringe_in.grid(row=i+3, column=2)
            entry_conc.grid(row=i+3, column=3)

        save_button = tk.Button(self.reagents_tab, text="Save reagents", command=self.save_reagents, bg = '#c7bffb', fg = 'black', font = ("Helvetica", 12))
        save_button.grid(row=num_rows+4, columnspan=3)


    def create_experiment(self):
        num_entries = int(self.num_syringes_combobox.get())
        label_time_of_change = tk.Label(self.experiment_tab, text="Time", font=("Helvetica", 12), bg='#e0dcfd')
        label_time_of_change.grid(row=3, column=0)
        label_start = tk.Label(self.experiment_tab, text="0", font=("Helvetica", 12), bg='#e0dcfd')
        label_start.grid(row=4, column=0)
        for i in range(num_entries):
            label_syringe_flow = tk.Label(self.experiment_tab, text=f"Syringe {i+1}", font=("Helvetica", 12), bg='#e0dcfd')
            entry_syringe_flow = tk.Entry(self.experiment_tab, font=("Helvetica", 12))
            label_syringe_flow.grid(row=3, column=i+1)
            entry_syringe_flow.grid(row=4, column=i+1)

    def add_reagent_names(self):
        num_rows = int(self.num_reagents_combobox.get()) + int(self.num_add_reagents_combobox.get())
        num_syr = int(self.num_syringes_combobox.get())
        label_reagent_name = tk.Label(self.reagents_tab, text=f"Reagent name", font=("Helvetica", 12), bg='#e0dcfd')
        label_syringe_in = tk.Label(self.reagents_tab, text=f"Syringe", font=("Helvetica", 12), bg='#e0dcfd')
        label_concentration = tk.Label(self.reagents_tab, text=f"Concentration (mM)", font=("Helvetica", 12), bg='#e0dcfd')
        label_reagent_name.grid(row=2, column=1)
        label_syringe_in.grid(row=2, column=2)
        label_concentration.grid(row=2, column=3)

        for i in range(num_rows):
            label_reagent = tk.Label(self.reagents_tab, text=f"{i+1}", font=("Helvetica", 12), bg='#e0dcfd')
            entry_reagent_name = tk.Entry(self.reagents_tab, font=("Helvetica", 12))
            entry_syringe_in = ttk.Combobox(self.reagents_tab, values=list(range(1, int(num_syr)+1)), font=("Helvetica", 12))
            entry_conc = tk.Entry(self.reagents_tab, font=("Helvetica", 12))
            label_reagent.grid(row=i+3, column=0)
            entry_reagent_name.grid(row=i+3, column=1)
            entry_syringe_in.grid(row=i+3, column=2)
            entry_conc.grid(row=i+3, column=3)

        save_button = tk.Button(self.reagents_tab, text="Save reagents", command=self.save_reagents, bg = '#c7bffb', fg = 'black', font = ("Helvetica", 12))
        save_button.grid(row=num_rows+4, columnspan=3)

    def save_reagents(self):
        num_reagents = int(self.num_reagents_combobox.get())
        num_add = int(self.num_add_reagents_combobox.get())
        filename = self.exp_name_entry.get() + '.xlsx'
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.create_sheet("reagents")
        headers = ["Reagent Name", "Syringe", "Concentration (mM)"]
        worksheet.append(headers)
        for i in range(num_reagents + num_add):
            reagent_name = self.reagents_tab.grid_slaves(row=i+3, column=1)[0].get()
            syringe_in = self.reagents_tab.grid_slaves(row=i+3, column=2)[0].get()
            concentration = float(self.reagents_tab.grid_slaves(row=i+3, column=3)[0].get())
            row_data = [reagent_name, syringe_in, concentration]
            worksheet.append(row_data)
        workbook.save(filename)

    def add_rows(self):
        num_entries = int(self.num_syringes_combobox.get())
        self.num_rows += 1  
        for i in range(num_entries+1):
            new_change_entry = tk.Entry(self.experiment_tab, font=("Helvetica", 12))
            new_change_entry.grid(row=self.num_rows, column=i)
        
    
    def save_changes(self):
        num_entries = int(self.num_syringes_combobox.get())
        filename = self.exp_name_entry.get() + '.xlsx'
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.create_sheet("experiment")
        headers = ["Time (min)"]
        for i in range(num_entries):
            headers.append(f"Syringe {i+1}")
        worksheet.append(headers)
        row_data = ["0"] 
        for i in range(num_entries):
            row_data.append(float(self.experiment_tab.grid_slaves(row=4, column=i+1)[0].get()))   
        worksheet.append(row_data)
        for row in range(5, self.num_rows + 1): 
            row_data=[]
            for i in range(num_entries+1):
                row_data.append(float(self.experiment_tab.grid_slaves(row=row, column=i)[0].get()))  
            worksheet.append(row_data)
        worksheet.cell(row=self.num_rows, column=1, value="Total time (min)")
        worksheet.cell(row=self.num_rows, column=2, value=float(self.total_time_entry.get()))
        workbook.save(filename)
   
    def calc(self):
        filename = self.exp_name_entry.get() + '.xlsx'
        filename_2 = self.exp_name_entry.get() + '.xlsx'
        try:
            calculations_CSTR(filename, filename_2)
            messagebox.showinfo("Success", "Calculation is complete.")
        except Exception:
            messagebox.showerror("Error", "An error occurred. Check that you filled all the boxes correctly.")

