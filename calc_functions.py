import openpyxl
from openpyxl.styles import Alignment
import math
from openpyxl.chart import ScatterChart, Reference, Series


def calculations_CSTR(filename, filename_2):

    workbook = openpyxl.load_workbook(filename)
    first_sheet = workbook.worksheets[0]
    second_sheet = workbook.worksheets[1]
    third_sheet = workbook.worksheets[2]
    
    format_cells(third_sheet)
    format_cells(second_sheet)
    format_cells(first_sheet)

    total_flow_calc(third_sheet)

    unique_reagents, duplicate_reagents = categorize_reagents(second_sheet)
    sheets_creator(workbook, first_sheet, second_sheet, third_sheet, unique_reagents, duplicate_reagents)

    workbook.save(filename_2)

def format_cells(sheet):

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

def total_flow_calc(third_sheet):

    col_for_tot_flow = third_sheet.max_column + 1
    third_sheet.cell(row=1, column=col_for_tot_flow, value="total flow")
    for row in range(2, third_sheet.max_row - 1):
        total_flow = sum([third_sheet.cell(row=row, column=col).value for col in range(2, third_sheet.max_column)])
        third_sheet.cell(row=row, column=col_for_tot_flow, value=total_flow)

def categorize_reagents(second_sheet):

    reagent_counts = {}
    for row in range(2, second_sheet.max_row + 1):
        reagent = second_sheet.cell(row=row, column=1).value
        if reagent in reagent_counts:
            reagent_counts[reagent] += 1
        else:
            reagent_counts[reagent] = 1

    unique_reagents = [reagent for reagent, count in reagent_counts.items() if count == 1]
    duplicate_reagents = [reagent for reagent, count in reagent_counts.items() if count > 1]

    return unique_reagents, duplicate_reagents

def sheets_creator(workbook, first_sheet, second_sheet, third_sheet, unique_reagents, duplicate_reagents):

    for row in range(2, second_sheet.max_row + 1):
        reagent = second_sheet.cell(row=row, column=1).value

        if reagent in unique_reagents:
            new_sheet = workbook.create_sheet(title=reagent)
            syringe = second_sheet.cell(row=row, column=2).value
            data_to_copy = int(syringe) + 1
            copypasting(first_sheet, second_sheet, third_sheet, new_sheet, data_to_copy, row)
            calc_combo(new_sheet)
            format_cells(new_sheet)

        elif reagent in duplicate_reagents:
            if reagent not in workbook.sheetnames:
                new_sheet = workbook.create_sheet(title=reagent)
                first_syringe = second_sheet.cell(row=row, column=2).value
                data_to_copy = int(first_syringe) + 1
                copypasting(first_sheet, second_sheet, third_sheet, new_sheet, data_to_copy, row)
                copypasting_add(second_sheet, third_sheet, new_sheet, reagent, row)
                calc_combo_for_two(new_sheet)
                format_cells(new_sheet)

def copypasting(first_sheet, second_sheet, third_sheet, new_sheet, data_to_copy, row):

    for col in range(1, 4):
        new_sheet.cell(row=1, column=col, value=second_sheet.cell(row=1, column=col).value)
        new_sheet.cell(row=2, column=col, value=second_sheet.cell(row=row, column=col).value)

    for col in range(1, 3):
        new_sheet.cell(row=1, column=col+6, value=third_sheet.cell(row=third_sheet.max_row, column=col).value)
        
    for row in range(1, third_sheet.max_row-1):
        new_sheet.cell(row=row+4, column=1, value=third_sheet.cell(row=row, column=1).value)
        new_sheet.cell(row=row+4, column=2, value=third_sheet.cell(row=row, column=data_to_copy).value) 
        new_sheet.cell(row=row+4, column=3, value=third_sheet.cell(row=row, column=third_sheet.max_column).value) 

    for col in range(1, 3):
        new_sheet.cell(row=2, column=col+6, value=first_sheet.cell(row=1, column=col).value) 

def copypasting_add(second_sheet, third_sheet, new_sheet, reagent, row):
     
     for duplicate_row in range(2, second_sheet.max_row + 1):
        if second_sheet.cell(row=duplicate_row, column=1).value == reagent and duplicate_row != row:
            second_syringe = int(second_sheet.cell(row=duplicate_row, column=2).value)
            data_2_to_copy = int(second_syringe)+1
            for col in range(1, 4):
                new_sheet.cell(row=3, column=col, value=second_sheet.cell(row=duplicate_row, column=col).value)

            for r in range(1, third_sheet.max_row - 1):
                new_sheet.cell(row=r + 4, column=3, value=third_sheet.cell(row=r, column=data_2_to_copy).value)
                new_sheet.cell(row=r + 4, column=4, value=third_sheet.cell(row=r, column=third_sheet.max_column).value)

def values_in_memory(sheet):

    times = []
    substance_flows = []
    total_flows = []
    tot_time = sheet['H1'].value

    for r in range(6, sheet.max_row+1):
        time = float(sheet.cell(row=r, column=1).value)
        s_f = float(sheet.cell(row=r, column=2).value)/60
        t_f = float(sheet.cell(row=r, column=3).value)/60
        times.append(time)
        substance_flows.append(s_f)
        total_flows.append(t_f)

    return  times, substance_flows, total_flows, tot_time

def values_in_memory_for_two(sheet):

    times = []
    substance_flows_1 = []
    substance_flows_2 = []
    total_flows = []
    tot_time = sheet['H1'].value

    for r in range(6, sheet.max_row+1):
        time = float(sheet.cell(row=r, column=1).value)
        s_f_1 = float(sheet.cell(row=r, column=2).value)/60
        s_f_2 = float(sheet.cell(row=r, column=3).value)/60
        t_f = float(sheet.cell(row=r, column=4).value)/60
        times.append(time)
        substance_flows_1.append(s_f_1)
        substance_flows_2.append(s_f_2)
        total_flows.append(t_f)

    return  times, substance_flows_1, substance_flows_2, total_flows, tot_time

def input_concentration_over_time(times, substance_flows, total_flows, tot_time, initial_concentration):

    concentration_table = []
    current_time=0

    for i in range(len(times)):
        if i < len(times) - 1:
            next_time = times[i + 1]
        else:
            next_time = tot_time
        
        while current_time < next_time:
            current_concentration = initial_concentration * (substance_flows[i] / total_flows[i])
            concentration_table .append(current_concentration)
            current_time += 1
       
    return concentration_table

def input_conc_for_two(tot_time, concentration_table_1, concentration_table_2):
    concentration_table_tot = []
    for i in range(0, tot_time):
        conc = float(concentration_table_1[i]) + float(concentration_table_2[i])
        concentration_table_tot.append(conc)

    return concentration_table_tot

def flow_rates(times, total_flows, tot_time):

    tot_flows_table = []
    current_time=0
    for i in range(len(times)):
        if i < len(times) - 1:
            next_time = times[i + 1]
        else:
            next_time = tot_time
        
        while current_time < next_time:
            current_flow=total_flows[i]
            tot_flows_table.append(current_flow)
            current_time += 1

    return tot_flows_table

def add_time_concentration_table(sheet):

    sheet.cell(row=1, column=10, value="time (min)")
    sheet.cell(row=1, column=11, value="concentration of reagent (mM)")
    max_time = int(sheet.cell(row=1, column=8).value)
    for i in range(0, max_time):
        sheet.cell(row=i + 2, column=10, value=i)

def starting_conc(sheet):

    flow=float(sheet.cell(row=6, column=2).value)
    tot_flow=float(sheet.cell(row=6, column=3).value)
    conc_in_syr=float(sheet.cell(row=2, column=3).value)
    starting_conc=(flow/tot_flow)*conc_in_syr
    sheet.cell(row=2, column=11, value=starting_conc)

def starting_conc_for_two_syr(sheet):

    flow_1=float(sheet.cell(row=6, column=2).value)
    flow_2=float(sheet.cell(row=6, column=3).value)
    tot_flow=float(sheet.cell(row=6, column=4).value)
    conc_in_syr_1=float(sheet.cell(row=2, column=3).value)
    conc_in_syr_2=float(sheet.cell(row=3, column=3).value)
    sum=(flow_1*conc_in_syr_1)+(flow_2*conc_in_syr_2)
    starting_conc=sum/tot_flow
    sheet.cell(row=2, column=11, value=starting_conc)

def calculate_table(sheet, tot_flows_table, tot_time):

   volume = sheet['H2'].value

   for row in range(3, tot_time+2):
        for flow_rate in tot_flows_table:
            previous_concentration = sheet.cell(row=row-1, column=11).value 
            flow_concentration = sheet.cell(row=row, column=12).value
            current_concentration = flow_concentration + (previous_concentration - flow_concentration) * math.exp(-flow_rate / volume)
            sheet.cell(row=row, column=11, value=current_concentration)

def create_scatter_plot(sheet):

    chart = ScatterChart()
    chart.title = "Concentration over time"
    chart.x_axis.title = "Time (min)"
    chart.y_axis.title = "Concentration (mM)"
    chart.legend = None

    x_values = Reference(sheet, min_col=10, min_row=2, max_row=sheet.max_row-1)
    y_values = Reference(sheet, min_col=11, min_row=2, max_row=sheet.max_row-1)

    series = Series(y_values, x_values, title_from_data=False)
    chart.series.append(series)

    sheet.add_chart(chart, "N1")  

def calc_combo(sheet):

    times, substance_flows, total_flows, tot_time = values_in_memory(sheet)
    initial_concentration = sheet['C2'].value
    concentration_table = input_concentration_over_time(times, substance_flows, total_flows, tot_time, initial_concentration)
    tot_flows_table = flow_rates(times, total_flows, tot_time)
    add_time_concentration_table(sheet)
    starting_conc(sheet)
    
    sheet.cell(row=1, column=13).value = "Total flow (microliteres/min)"
    for minute, tot_flow in enumerate(tot_flows_table):
        sheet.cell(row=minute + 2, column=13).value = tot_flow
    sheet.cell(row=1, column=12).value = "Input concentration(mM)"
    for minute, concentration in enumerate(concentration_table):
        sheet.cell(row=minute + 2, column=12).value = concentration
    calculate_table(sheet, tot_flows_table, tot_time)
    create_scatter_plot(sheet)

def calc_combo_for_two(sheet):

    times, substance_flows_1, substance_flows_2, total_flows, tot_time = values_in_memory_for_two(sheet)
    initial_concentration_1 = sheet['C2'].value
    concentration_table_1 = input_concentration_over_time(times, substance_flows_1, total_flows, tot_time, initial_concentration_1)
    initial_concentration_2 = sheet['C3'].value
    concentration_table_2 = input_concentration_over_time(times, substance_flows_2, total_flows, tot_time, initial_concentration_2)
    concentration_table_total = input_conc_for_two(tot_time, concentration_table_1, concentration_table_2)
    tot_flows_table = flow_rates(times, total_flows, tot_time)
    add_time_concentration_table(sheet)
    starting_conc_for_two_syr(sheet)
    
    sheet.cell(row=1, column=13).value = "Total flow (microliteres/min)"
    for minute, tot_flow in enumerate(tot_flows_table):
        sheet.cell(row=minute + 2, column=13).value = tot_flow
    
    sheet.cell(row=1, column=12).value = "Input concentration(mM)"
    for minute, concentration in enumerate(concentration_table_total):
        sheet.cell(row=minute + 2, column=12).value = concentration

    calculate_table(sheet, tot_flows_table, tot_time)
    create_scatter_plot(sheet)   
