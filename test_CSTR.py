import pytest
import openpyxl
import tempfile
import shutil
from calc_functions import calculations_CSTR

@pytest.fixture
def setup_test_environment():
    temp_dir = tempfile.mkdtemp()
    output_file = f"{temp_dir}/test_output.xlsx"
    yield output_file
    shutil.rmtree(temp_dir)

def test_calculations_CSTR(setup_test_environment):
    input_file = 'TEST.xlsx'
    sample_file = 'SAMPLE.xlsx'
    output_file = setup_test_environment

    calculations_CSTR(input_file, output_file)

    workbook_output = openpyxl.load_workbook(output_file)
    workbook_sample = openpyxl.load_workbook(sample_file)

    def compare_workbooks(wb1, wb2):
        if wb1.sheetnames != wb2.sheetnames:
            return False
        for sheetname in wb1.sheetnames:
            sheet1 = wb1[sheetname]
            sheet2 = wb2[sheetname]
            if sheet1.dimensions != sheet2.dimensions:
                return False
            for row in sheet1.iter_rows():
                for cell in row:
                    if cell.value != sheet2[cell.coordinate].value:
                        return False
        return True

    assert compare_workbooks(workbook_output, workbook_sample), "The output file does not match the sample file"